import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
import logging

logging.basicConfig(filename='test_execution.log', level=logging.INFO)

# Page object model for the login page
class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "txtUsername"))
        )
        self.password_field = driver.find_element(By.ID, "txtPassword")
        self.login_button = driver.find_element(By.ID, "btnLogin")

    def login(self, username, password):
        self.username_field.send_keys(username)
        self.password_field.send_keys(password)
        self.login_button.click()

# Test function using data-driven approach
@pytest.mark.parametrize("data", load_workbook("test_data.xlsx").active.iter_rows(min_row=2))
def test_login(data):
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

    login_page = LoginPage(driver)
    login_page.login(data[1].value, data[2].value)

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "welcome"))
        )
        data[6].value = "Passed"
    except:
        logging.error("Login failed for user: {}".format(data[1].value))
        data[6].value = "Failed"

    finally:
        workbook = load_workbook("test_data.xlsx")
        worksheet = workbook.active
        worksheet.append(data)
        workbook.save("test_data.xlsx")
        driver.quit()
